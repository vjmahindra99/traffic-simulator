# Export stats from simulator to excel file
import os
from . import settings
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment

# Time conversion function (seconds to minutes/seconds)
def sec_to_min_sec(seconds):

    minutes = int(seconds // 60)
    secs = int(seconds % 60)
    return "{:02d}:{:02d}".format(minutes, secs)

# Export stats function
def export_stats_to_xlsx(filename="smart_stats.xlsx"):
    
    # Path to save excel file
    filepath = os.path.join(os.path.dirname(__file__), "results", filename)

    # Direction order based on your settings
    dirs = ["right", "down", "left", "up"]  # 0: right, 1: down, 2: left, 3: up
    crossed = [settings.vehicles[d]["crossed"] for d in dirs]

    # Raw sums & counts
    lane_wait_sum = [sec_to_min_sec(v) for v in settings.lane_wait_sum]
    lane_wait_count = settings.lane_wait_count

    lane_wait_before_green_sum = [sec_to_min_sec(v) for v in settings.lane_wait_before_green_sum]
    lane_wait_before_green_count = settings.lane_wait_before_green_count

    queue_at_green_sum = [sec_to_min_sec(v) for v in settings.queue_at_green_sum]
    queue_at_green_count = settings.queue_at_green_count

    total_passed = sum(crossed)
    time_elapsed = sec_to_min_sec(settings.time_elapsed)

    # Ambulance priority stats
    priority_total = settings.ambulance_priority_total
    priority_lane = settings.ambulance_priority_per_lane

    # Create workbook & sheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Smart Stats"

    # Title
    title_text = "Smart Traffic Light Statistics"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
    title_cell = ws.cell(row=1, column=1, value=title_text)

    title_cell.font = Font(size=14, bold=True, color="FFFFFF")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    title_cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid") # dark red top bar
    row = 3 # Leave a blank row

    # Table styles
    header_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid") # red colour
    header_font = Font(color="FFFFFF", bold=True)
    row_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

    thin_side = Side(style="thin")
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    center_align = Alignment(horizontal="center", vertical="center")

    def style_header_range(row_idx, start_col, end_col):
        for col in range(start_col, end_col + 1):
            cell = ws.cell(row=row_idx, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = thin_border

    def style_data_range(row_idx, start_col, end_col, shaded=False):
        for col in range(start_col, end_col + 1):
            cell = ws.cell(row=row_idx, column=col)
            if shaded:
                cell.fill = row_fill
            cell.alignment = center_align
            cell.border = thin_border

    # Meta info
    ws.cell(row=row, column=1, value="Metric")
    ws.cell(row=row, column=2, value="Value")
    style_header_range(row, 1, 2)
    row += 1

    ws.cell(row=row, column=1, value="Time elapsed (mm:ss)")
    ws.cell(row=row, column=2, value=time_elapsed)
    style_data_range(row, 1, 2, shaded=True)
    row += 1

    ws.cell(row=row, column=1, value="Total vehicles passed")
    ws.cell(row=row, column=2, value=total_passed)
    style_data_range(row, 1, 2)
    row += 1

    ws.cell(row=row, column=1, value="Total ambulance-priority events")
    ws.cell(row=row, column=2, value=priority_total)
    style_data_range(row, 1, 2, shaded=True)
    row += 2 # blank line

    # Per-direction vehicle count
    headers = ["Direction", "Right", "Down", "Left", "Up"]
    for col, h in enumerate(headers, start=1):
        ws.cell(row=row, column=col, value=h)
    style_header_range(row, 1, 5)
    row += 1

    ws.cell(row=row, column=1, value="Vehicles crossed")
    for i, val in enumerate(crossed, start=2):
        ws.cell(row=row, column=i, value=val)
    style_data_range(row, 1, 5, shaded=True)
    row += 2 # blank line

    # Ambulance priority per lane
    ws.cell(row=row, column=1, value="Ambulance priority events per lane")
    ws.cell(row=row, column=2, value="Right")
    ws.cell(row=row, column=3, value="Down")
    ws.cell(row=row, column=4, value="Left")
    ws.cell(row=row, column=5, value="Up")
    style_header_range(row, 1, 5)
    row += 1

    ws.cell(row=row, column=1, value="Times lane was chosen due to ambulance")
    for i, val in enumerate(priority_lane, start=2):
        ws.cell(row=row, column=i, value=val)
    style_data_range(row, 1, 5, shaded=True)
    row += 2 # blank line

    # Lane wait raw sums
    ws.cell(row=row, column=1, value="Metric")
    ws.cell(row=row, column=2, value="Dir_0 (Right)")
    ws.cell(row=row, column=3, value="Dir_1 (Down)")
    ws.cell(row=row, column=4, value="Dir_2 (Left)")
    ws.cell(row=row, column=5, value="Dir_3 (Up)")
    style_header_range(row, 1, 5)
    row += 1

    # Total accumulated waiting time for individual vehicles
    ws.cell(row=row, column=1, value="Total wait time per lane (mm:ss)")
    for i, val in enumerate(lane_wait_sum, start=2):
        ws.cell(row=row, column=i, value=val)
    style_data_range(row, 1, 5, shaded=True)
    row += 1

    # Number of vehicles used to compute the total
    ws.cell(row=row, column=1, value="Total vehicle count wait time (per lane)")
    for i, val in enumerate(lane_wait_count, start=2):
        ws.cell(row=row, column=i, value=val)
    style_data_range(row, 1, 5)
    row += 2

    # Total accumulated time the lane spent waiting between green cycles
    ws.cell(row=row, column=1, value="Green phase interval sum (mm:ss)")
    for i, val in enumerate(lane_wait_before_green_sum, start=2):
        ws.cell(row=row, column=i, value=val)
    style_data_range(row, 1, 5, shaded=True)
    row += 1

    # Number of green signal cycles
    ws.cell(row=row, column=1, value="Green light cycle count")
    for i, val in enumerate(lane_wait_before_green_count, start=2):
        ws.cell(row=row, column=i, value=val)
    style_data_range(row, 1, 5)
    row += 2

    # Total number of vehicles that were already waiting in the lane the moment it turned green
    ws.cell(row=row, column=1, value="Vehicles at lane when green phase")
    for i, val in enumerate(queue_at_green_sum, start=2):
        ws.cell(row=row, column=i, value=val)
    style_data_range(row, 1, 5, shaded=True)
    row += 1

    # How many times this lane turned green
    ws.cell(row=row, column=1, value="Green phase count per lane")
    for i, val in enumerate(queue_at_green_count, start=2):
        ws.cell(row=row, column=i, value=val)
    style_data_range(row, 1, 5)

    # Auto-fit column widths
    for col_idx in range(1, ws.max_column + 1):
        max_len = 0
        col_letter = get_column_letter(col_idx)
        for row_idx in range(1, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value is not None:
                length = len(str(cell.value))
                if length > max_len:
                    max_len = length
        ws.column_dimensions[col_letter].width = max_len + 2

    # Save workbook
    wb.save(filepath)
    print("[Smart] stats exported to", filepath)