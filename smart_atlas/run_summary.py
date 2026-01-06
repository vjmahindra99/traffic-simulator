# Summary stats for per finished simulation
import os
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from . import settings
from .export_stats import sec_to_min_sec

# Excel file for summarising simulation run
def append_run_summary(filename="smart_runs_summary.xlsx"):

    # Path to excel file
    filepath = os.path.join(os.path.dirname(__file__), "results", filename)

    # Create workbook
    if os.path.exists(filepath):
        wb = load_workbook(filepath)
        ws = wb.active
        new_file = False
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Run Summary"
        new_file = True

    # Font styling
    header_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    row_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

    thin_side = Side(style="thin")
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    center_align = Alignment(horizontal="center", vertical="center")

    def style_header_range(row_idx, start_col, end_col):
        for col in range(start_col, end_col + 1):
            cell = ws.cell(row=row_idx, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = thin_border

    def style_data_range(row_idx, start_col, end_col, shaded=False):
        for col in range(start_col, end_col + 1):
            cell = ws.cell(row=row_idx, column=col)
            if shaded:
                cell.fill = row_fill
            cell.alignment = center_align
            cell.border = thin_border

    # Helper: "mm:ss" -> seconds
    def mmss_to_seconds(mmss):
        try:
            parts = str(mmss).strip().split(":")
            if len(parts) != 2:
                return 0
            m = int(parts[0])
            s = int(parts[1])
            return m * 60 + s
        except Exception:
            return 0

    # Helper: seconds -> "mm:ss"
    def seconds_to_mmss(total_seconds):
        try:
            total_seconds = int(round(float(total_seconds)))
        except Exception:
            total_seconds = 0
        m = total_seconds // 60
        s = total_seconds % 60
        return "{:02d}:{:02d}".format(m, s)

    # Workbook layout
    title_row = 1
    blank1_row = 2
    summary_row = 3
    blank2_row = 4
    header_row = 5
    first_run_row = 6

    # Title and headers
    if new_file:

        title_text = "Smart Traffic Run Summary"
        ws.merge_cells(start_row=title_row, start_column=1, end_row=title_row, end_column=4)
        title_cell = ws.cell(row=title_row, column=1, value=title_text)
        title_cell.font = Font(size=14, bold=True, color="FFFFFF")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        title_cell.fill = header_fill

        ws.cell(row=blank1_row, column=1, value="")
        ws.cell(row=blank2_row, column=1, value="")

        # Summary 2 row
        ws.cell(row=summary_row, column=1, value="Total runs")
        ws.cell(row=summary_row, column=2, value=0)
        ws.cell(row=summary_row, column=3, value="Avg time (mm:ss)")
        ws.cell(row=summary_row, column=4, value="00:00")

        # Summary 2 style
        style_header_range(summary_row, 1, 1)
        style_header_range(summary_row, 3, 3)
        style_data_range(summary_row, 2, 2, shaded=True)
        style_data_range(summary_row, 4, 4, shaded=True)

        # Header row
        ws.cell(row=header_row, column=1, value="Simulation Run")
        ws.cell(row=header_row, column=2, value="Time to complete (mm:ss)")
        ws.cell(row=header_row, column=3, value="Total vehicles passed")
        style_header_range(header_row, 1, 3)

    # Compute summary for each run
    dirs = ["right", "down", "left", "up"]
    crossed = [settings.vehicles[d]["crossed"] for d in dirs]
    total_passed = sum(crossed)
    time_elapsed_str = sec_to_min_sec(settings.time_elapsed)
    
    # Append new run
    next_row = ws.max_row + 1
    run_number = next_row - (first_run_row - 1)

    ws.cell(row=next_row, column=1, value=run_number)
    ws.cell(row=next_row, column=2, value=time_elapsed_str)
    ws.cell(row=next_row, column=3, value=total_passed)
    style_data_range(next_row, 1, 3, shaded=(run_number % 2 == 1))

    # Update running average time & total runs for summary 2
    last_run_row = ws.max_row

    times_sec = []
    for r in range(first_run_row, last_run_row + 1):
        val = ws.cell(row=r, column=2).value
        if val is None:
            continue
        sec = mmss_to_seconds(val)
        if sec > 0:
            times_sec.append(sec)

    total_runs = len(times_sec)
    if total_runs > 0:
        avg_sec = float(sum(times_sec)) / float(total_runs)
        avg_mmss = seconds_to_mmss(avg_sec)
    else:
        avg_mmss = "00:00"

    ws.cell(row=summary_row, column=2, value=total_runs)
    ws.cell(row=summary_row, column=4, value=avg_mmss)

    # Auto-fit column widths
    for col_idx in range(1, ws.max_column + 1):
        max_len = 0
        col_letter = get_column_letter(col_idx)
        for row_idx in range(1, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value is not None:
                length = len(str(cell.value))
                if length > max_len:
                    max_len = length
        ws.column_dimensions[col_letter].width = max_len + 2

    # Save workbook
    wb.save(filepath)
    print("[SMART] run summary appended to", filepath, "| runs =", total_runs, "| avg time =", avg_mmss)